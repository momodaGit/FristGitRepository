from selenium import webdriver
import time
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
driver = webdriver.Chrome(r'd:\chromedriver.exe')

#跳转路径配置
URL = 'http://bug.gobestsoft.com:81/zentao/user-login-L3plbnRhby8=.html'#首页路径

URL1 = 'http://bug.gobestsoft.com:81/zentao/project-bug-133.html'#反复提交bug定位界面

#生成BUG数组
BUGdata = [] #bug名称列表
BUGperson = []
#读取的行数最大值（每次需要调整）
max_row_num = 3

#读取文件，获取EXCEL表格中第一列的信息，bug名字
def Read_file():
    #读取路径
    book = load_workbook(filename=r"C:\Users\User\Desktop\缺陷表导入.xlsx")
    #读取名字为Sheet1的表
    sheet = book.get_sheet_by_name("Sheet1")
    #用于存储数据的数组
    row_num = 1
    while row_num <= max_row_num:
        #将表中第一列的max行数据之前（包括max）写入data数组中
        BUGdata.append(sheet.cell(row=row_num, column=1).value)
        row_num = row_num + 1
    return BUGdata

#读取文件，获取EXCEL表格中第一列的信息，bug指派人
def Read_file1():
    #读取路径
    book = load_workbook(filename=r"C:\Users\User\Desktop\缺陷表导入.xlsx")
    #读取名字为Sheet1的表
    sheet = book.get_sheet_by_name("Sheet1")
    #用于存储数据的数组
    row_num = 1
    while row_num <= max_row_num:
        #将表中第一列的max行数据之前（包括max）写入BUGperson数组中
        BUGperson.append(sheet.cell(row=row_num, column=2).value)
        row_num = row_num + 1
    return BUGperson

#浏览器全屏操作
def interface():
    driver.maximize_window()

#直接访问网站地址
def open():
    driver.get(URL)

#单网页登录操作，输入账号/密码
def login():
    time.sleep(1)
    driver.find_element_by_xpath("//input[@name='account']").send_keys('zhangjiawei')
    time.sleep(1)
    driver.find_element_by_xpath("//input[@name='password']").send_keys('zhangjiawei')
    time.sleep(1)
    driver.find_element_by_xpath("//button[@type='submit']").click()
    time.sleep(2)

#进入太保项目bug管理界面点击提交新bug
def Enter_Into():
    #进入项目
    driver.find_element_by_link_text('项目').click()
    time.sleep(1)
    driver.find_element_by_xpath("//button[@id='currentItem']").click()
    #进入太平洋保险APP这一列---定位作用
    driver.find_element_by_xpath("//input[@class='form-control search-input empty']").send_keys('太平洋保险集团党工团APP')
    driver.find_element_by_xpath("//a[@title='太平洋保险集团党工团APP']").click()
    time.sleep(1)


#多次提交BUG提交操作
def Submit_Again():
    i = 0
    for j in range(len(BUGdata)):    
        Logging_data = BUGdata[i]
        Logging_person = BUGperson[i]
        #点击提交BUG按钮
        driver.find_element_by_xpath("//a[@href='/zentao/project-bug-133.html']").click()
        driver.find_element_by_xpath("//a[@href='/zentao/bug-create-39-0-projectID=133.html']").click()
        #填入BUG指派人
        driver.find_element_by_xpath("//a[@class='chosen-single chosen-single-with-deselect chosen-default']/span[@title=' ']").click().send_keys(Logging_person).send_keys(Keys.ENTER)
        #填入BUG名称信息
        driver.find_element_by_xpath("//input[@autocomplete='off' and @name='title']").send_keys(Logging_data)
        
        #点击提交当前BUG
        #driver.find_element_by_xpath("//button[@id='submit' and @data-loading='稍候...' and@class='btn btn-wide btn-primary']").click()
        time.sleep(1)
        #再次定位提交新BUG界面
        #driver.get(URL1)
        i = i + 1
        time.sleep(2)

#单网页直接点击登录操作，反复提交BUG
def GO():
    interface()#全屏操作
    open()#打开首页
    login()#首页输入账号密码
    Enter_Into()#进入到党工团项目
    Read_file()#读取Excel文件内容生成数组bug名称
    Read_file1()#读取Excel文件内容生成数组bug指派人
    print(BUGperson)
    print()
    print(BUGdata)
    Submit_Again()#反复提交BUG操作

#调用方法    
GO()




