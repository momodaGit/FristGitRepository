'''
BUG = ['缺陷一','缺陷二','缺陷三','缺陷四']







i = 0
for j in range(len(BUG)):    
    Logging_data = BUG[i]
    print(BUG[i])
    i+=1'''
max_row_num = 3
from openpyxl import load_workbook
BUGdata= []

def Read_file():
    #读取路径
    book = load_workbook(filename=r"C:\Users\User\Desktop\缺陷表导入.xlsx")
    #读取名字为Sheet1的表
    sheet = book.get_sheet_by_name("Sheet1")
    #用于存储数据的数组
    row_num = 1
    while row_num <= max_row_num:
        #将表中第一列的max行数据之前（包括max）写入data数组中
        BUGdata.append(sheet.cell(row=row_num, column=1).value)
        row_num = row_num + 1
    return BUGdata

Read_file()
print(BUGdata)