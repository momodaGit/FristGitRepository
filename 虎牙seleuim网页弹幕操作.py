

from selenium import webdriver
import time
from openpyxl import load_workbook
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
driver = webdriver.Chrome(r'd:\chromedriver.exe')

#虎牙账号配置用之前需要填写‘’中内容
Account = ''

#虎牙密码配置用之前需要填写‘’中内容
Password = ''

#跳转路径配置
URL = 'https://www.huya.com/?ptag=gouzai'#首页路径
account = Account
password = Password
max_row_num = 100
BUGdata = []
room = '血色-BB文'    #//p[@class='name']  这条定位的文本信息


#读取文件，获取EXCEL表格中第一列的信息
def Read_file():
    #读取路径
    book = load_workbook(filename=r"C:\Users\User\Desktop\弹幕.xlsx")
    #读取名字为Sheet1的表
    sheet = book.get_sheet_by_name("Sheet1")
    #用于存储数据的数组
    row_num = 1
    while row_num <= max_row_num:
        #将表中第一列的max行数据之前（包括max）写入data数组中
        BUGdata.append(sheet.cell(row=row_num, column=1).value)
        row_num = row_num + 1
    return BUGdata

#浏览器全屏操作
def interface():
    driver.maximize_window()

#直接访问网站地址点击我的订阅
def open():
    driver.get(URL)
    time.sleep(1)
    #driver.find_element_by_xpath("//button[@type='button'and@class='btn btn-default']").click()
    driver.find_element_by_xpath("//span[@class='title clickstat'and@id='nav-login']").click()
    time.sleep(1)
    #跳转到frame
    driver.switch_to.frame(driver.find_element_by_xpath("//iframe[@id='UDBSdkLgn_iframe']"))
    #输入账号密码
    time.sleep(1)
    driver.find_element_by_xpath("//input[@type='text'and@class='udb-input udb-input-account']").send_keys(account)
    time.sleep(1)
    driver.find_element_by_xpath("//input[@type='password'and@placeholder='密码']").send_keys(password)
    #点击登录按钮
    time.sleep(1)
    driver.find_element_by_xpath("//div[@id='login-btn']").click()
    time.sleep(1)
    #回到HTML
    driver.switch_to.default_content()
    #获取订阅
    List = driver.find_elements_by_xpath("//span[@class='title']")
    #鼠标移动到订阅上悬浮不点击操作
    ActionChains(driver).move_to_element(List[2]).perform()
    #列出所有订阅的信息列表
    List1 = driver.find_elements_by_xpath("//p[@class='name']")
    for i in List1:
        print(i.text)
        if i.text == room:
            break
    i.click()

#发送弹幕
def Hair_barrage():
    i = 0
    for j in range(len(BUGdata)):
        time.sleep(1)
        driver.execute_script("window,scrollTo(0,2500);")
        driver.find_element_by_xpath("//textarea[@rows='2']").send_keys(BUGdata[i])
        time.sleep(6)
        driver.find_element_by_xpath("//textarea[@rows='2']").send_keys(Keys.ENTER)
        i += 1
        

def Go_pruton():
    Read_file()
    interface()
    open()
    Hair_barrage()
    print(BUGdata)
Go_pruton()

